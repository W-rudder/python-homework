from typing_extensions import final
import matplotlib.pyplot as plt
import numpy as np
import os
from numpy.core.fromnumeric import argmax
from openpyxl import Workbook,load_workbook
from pyecharts.charts import Map
from pyecharts import options as ops


plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus'] = False


filepath=r"C:\Users\86186\Desktop\现代程序设计\co2_demo"

class NotNumError(ValueError):
    def __init__(self,year,province,industry,type) -> None:
        self.message={}
        self.message["年份"]=year
        self.message["省份"]=province
        self.message["厂商"]=industry
        self.message["排放类型"]=type

class data_analyze:

    def __init__(self,filepath) -> None:
        self.filelist=[]
        self.citylist=[]
        self.industrylist=[]
        self.emissionlist=[]
        files=os.listdir(filepath)
        self.files=files
        for file in files:
            wb=load_workbook(os.path.join(filepath,file),data_only=True)

            self.filelist.append(wb)
        
        

        ws1=self.filelist[0]["Sum"]
        colA=ws1["A"]
        for i in colA:
            if i.value != None:
                self.citylist.append(i.value)
        row1=ws1[1]
        for i in row1:
            if i.value != None:
                self.emissionlist.append(i.value)
        

        ws2=self.filelist[0]["Beijing"]
        colA_2=ws2["A"]
        for i in colA_2:
            if i.value != None:
                self.industrylist.append(i.value)
        self.industrylist=self.industrylist[1:]


    def time_analyze(self,cityname,emission_type):
        res=[]
        row_num=self.citylist.index(cityname)+2
        col_num=self.emissionlist.index(emission_type)+2
        year=1997
        for i in self.filelist:
            ws=i["Sum"]
            try:
                if ws.cell(row_num,col_num).value == None:
                    raise NotNumError(year,cityname,"NONE",emission_type)
            except NotNumError as notnum:
                print("内容为空，相关信息：")
                print(notnum.message)
                ws.cell(row_num,col_num).value = 0
                
            finally:
                res.append(ws.cell(row_num,col_num).value)
                year += 1
        print("{}的{}类型排放量的年分布为:".format(cityname,emission_type))
        count=1997
        for i in res:
            print("{}年\t{}百万吨".format(count,i))
            count += 1
        return res

    def room_analyze(self,year):
        res=[]
        mid=[]
        num=year-1997
        wb=self.filelist[num]
        ws=wb["Sum"]
        colB=ws["B"]
        for i in range(1,31):
            try:
                if colB[i].value == None:
                    raise NotNumError(year,self.citylist[i-1],"NONE","Total")
            except NotNumError as notnum:
                print("内容为空，相关信息：")
                print(notnum.message)
                colB[i].value.value = 0
            finally:
                mid.append(colB[i].value)
        for i in range(30):
            res.append((self.citylist[i],mid[i]))
        res=sorted(res,key=lambda x:x[1],reverse=True)
        print("{}年的CO2排放量从高到低分布情况：".format(year))
        for i in range(30):
            print("{:<20}{}百万吨".format(res[i][0],res[i][1]))
        return res,mid

    def scaling_analyze_emissiontype(self,year,city):
        res=[]
        num=year-1997
        wb=self.filelist[num]
        ws=wb["Sum"]
        row_num=self.citylist.index(city)+2
        row=ws[row_num]
        total=row[1].value
        for i in range(2,20):
            res.append(row[i].value)
        try:
            for i in range(len(res)):
                res[i]=res[i]/total
        except ZeroDivisionError:
            print("{}年{}的CO2排放总量为0\n行号：{}".format(year,city,row_num))
        else:
            print("{}年{}的CO2排放类型比例如下：".format(year,city))
            for i in range(18):
                print("{:<28}{:.2%}".format(self.emissionlist[i+1],res[i]))

    def scaling_analyze_industry(self,year,city,type):
        res=[]
        num=year-1997
        wb=self.filelist[num]
        emissionlist=self.emissionlist[1:19]
        emissionlist.append("Total")
        col_num=emissionlist.index(type)+2
        ws=wb[city]
        col=ws[chr(col_num+64)]
        for i in range(4,len(col)):
            try:
                if col[i].value == None:
                    raise NotNumError(year,city,self.industrylist[i-4],type)
            except NotNumError as notnum:
                print("内容为空，相关信息：")
                print(notnum.message)
                col[i].value = 0
            finally:
                res.append(col[i].value)
        wb.save(os.path.join(filepath,self.files[num]))
        print("{}年{}的{}类型CO2排放量的分布情况如下：".format(year,city,type))
        for i in range(47):
            print("{:<75}{}".format(self.industrylist[i],res[i]))

class Visualization(data_analyze):
    def __init__(self, filepath) -> None:
        super().__init__(filepath)

    def time_visualization(self,cityname,emission_type):
        res=Visualization.time_analyze(self,cityname,emission_type)
        y=np.array(res)
        x=np.arange(1997,2016)
        plt.plot(x,y)
        plt.xticks(x)
        plt.title("{}的{}类型CO2排放量年度变化".format(cityname,emission_type))
        plt.xlabel("年份")
        plt.ylabel("排放量（百万吨）")
        plt.show()
    
    def room_visualization(self,year):
        res,mid=Visualization.room_analyze(self,year)
        value=np.array(mid)
        city=["北京","天津","河北","山西","内蒙古","辽宁","吉林","黑龙江","上海","江苏","浙江","安徽","福建","江西","山东","河南","湖北","湖南","广东","广西","海南","重庆","四川","贵州","云南","陕西","甘肃","青海","宁夏","新疆"]
        sequence=list(zip(city,value))
        title="CO2排放量空间分布"
        map=Map(ops.InitOpts(width="1900px",height="800px",bg_color="pink",page_title="全国CO2排放量分布图"))
        map.add(str(year)+"年",sequence,maptype="china",is_map_symbol_show=False,label_opts=ops.LabelOpts(is_show=True))
        map.set_global_opts(title_opts=ops.TitleOpts(title),visualmap_opts=ops.VisualMapOpts(max_=np.max(value),min_=np.min(value),range_text=["高排放量","低排放量"],range_color=["lightskyblue","yellow","red"]))
        map.render('./test.html')
        
    def scaling_visualization(self,year):
        res,mid=Visualization.room_analyze(self,year)
        value=mid
        title=str(year)+"年CO2排放量比例"
        plt.pie(value,labels=self.citylist[:30],autopct="%1.2f%%",pctdistance=0.9)
        plt.title(title)
        plt.show()

        

t=Visualization(filepath)
t.scaling_visualization(1997)
